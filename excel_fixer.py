# excel_fixer.py - FIXED BUDGET INITIAL + NO DOUBLE COUNTING
import io
import re
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tempfile
import os

# ============================================================================
# MISTRAL CONFIGURATION
# ============================================================================
MISTRAL_API_KEY = "em5oqjSdA1Nus9iUpa1MNAJtQA4YfCtK"
MISTRAL_URL = "https://api.mistral.ai/v1/chat/completions"
MISTRAL_MODEL = "mistral-small-latest"

# ============================================================================
# ACCOUNT NAME SEARCH PATTERNS
# ============================================================================
ACCOUNT_SEARCH = {
    # REVENUES
    "transient revenue": 12,
    "revenus horaires": 12,
    "monthly revenues": 13,
    "revenus mensuels": 13,
    "other monthly revenue": 17,
    "car-wash revenue": 14,
    "revenus lave-auto": 14,
    "hotel revenue": 15,
    "revenus hotel": 15,
    "revenus hôtel": 15,
    "interests": 16,
    "intérêts": 16,
    "interets": 16,
    "miscellaneous": 17,
    "autres revenus": 17,
    "discount-gratuities - transient": 20,
    "gratuities - transient": 20,
    "discount-gratuities - monthly": 22,
    "gratuities - monthly": 22,
    # LABOUR
    "parking wages": 29,
    "salaire stationnement": 29,
    "other wages": 30,
    "salaire superviseur": 30,
    "training & recr.": 31,
    "formation & recrutement": 31,
    "uniformes": 32,
    "uniforms": 32,
    # MAINTENANCE
    "repair and maintenance": 37,
    "r&m - equipment": 37,
    "entretien équipement": 37,
    "entretien equipement": 37,
    "r&m - cleaning": 35,
    "nettoyage stationnement": 35,
    "r&m - general": 36,
    "entretien stationnement": 36,
    "r&m - signs": 38,
    "signalisation": 38,
    "r&m - lines": 39,
    "lignage": 39,
    "snow removal": 40,
    "déneigement": 40,
    "deneigement": 40,
    "parking supplies": 41,
    "fournitures stationnement": 41,
    "misc. re-billing": 42,
    "refacturations diverses": 42,
    # PUBLIC SERVICES
    "public services": 46,
    "services publics": 46,
    # OVERHEAD
    "office expenses": 49,
    "fournitures de bureau": 49,
    "telecommunication": 50,
    "télécommunication": 50,
    "télécommunications": 50,
    "rent": 51,
    "loyer": 51,
    "travel expenses": 52,
    "frais de déplacement": 52,
    "frais de deplacement": 52,
    "credit card fees": 53,
    "frais de cartes de crédit": 53,
    "frais de cartes de credit": 53,
    "bank fees": 54,
    "intérêts et frais de banque": 54,
    "interets et frais de banque": 54,
    "cash transportation fees": 55,
    "transport de fonds": 55,
    "claims": 56,
    "réclamations": 56,
    "reclamations": 56,
    "insurance & guarantee": 57,
    "assurances et cautionnement": 57,
    "tax & license": 58,
    "taxes et permis": 58,
    "professional services": 59,
    "comptabilité": 59,
    "comptabilite": 59,
    "equipment rent": 60,
    "location d'équipement": 60,
    "location d'equipement": 60,
    "ad. & promotion": 61,
    "publicité et promotion": 61,
    "publicite et promotion": 61,
    "percent management fee": 62,
    "honoraires de gestion en %": 62,
    "management fees (basic)": 63,
    "honoraires de gestion de base": 63,
    "incentives": 64,
    "incitatif annuel": 64,
    # OTHER EXPENSES
    "depreciation": 67,
    "amortissement": 67,
    "financial fees": 68,
    "intérêts sur emprunts": 68,
    "interets sur emprunts": 68,
    "security": 69,
    "sécurité": 69,
    "securite": 69,
    "co-ownership expenses": 70,
    "frais de copropriété": 70,
    "frais de copropriete": 70,
    "shuttle expenses": 71,
    "frais de navettes": 71,
    "computer services": 72,
    "services informatiques": 72,
    "bad debts": 73,
    "mauvaises créances": 73,
    "mauvaises creances": 73,
    "dues & subscription": 74,
    "cotisations": 74,
    "meal & entertainment": 76,
    "représentation repas": 76,
    "representation repas": 76,
}

# Keywords to EXCLUDE from matching (subtotals, headers, totals)
EXCLUDE_KEYWORDS = [
    "total", "parking revenue", "management revenue", "operation expenses",
    "operating expenses", "operation surplus", "other expenses",
    "total revenue", "total operation", "net income", "total other",
    "total frais", "total dépenses", "total depenses", "total des frais",
    "total des revenus", "bénéfice net", "benefice net",
    "revenus de stationnement", "dépenses", "depenses",
    "autres frais", "total autres", "total entretien",
    "total frais de personnel", "total services publics",
    "total frais généraux", "total autres dépenses",
]

# Revenue keywords for classification
REVENUE_KEYWORDS = [
    "revenue", "revenus", "revenu", "income", "interests", "intérêts",
    "car-wash", "lave-auto", "hotel", "miscellaneous", "autres",
    "monthly", "mensuels", "transient", "horaires", "journaliers",
    "violation", "shuttle", "navettes", "valet", "subvention"
]

EXPENSE_KEYWORDS = [
    "wages", "salaires", "salaire", "uniforms", "uniformes",
    "training", "formation", "recrutement", "supplies", "fournitures",
    "cleaning", "nettoyage", "equipment", "équipement", "equipement",
    "maintenance", "entretien", "repair", "réparation", "reparation",
    "signs", "signalisation", "lines", "lignage", "snow", "neige",
    "déneigement", "deneigement", "security", "sécurité", "securite",
    "rent", "loyer", "tax", "taxes", "permis", "license",
    "insurance", "assurances", "cautionnement", "claims", "réclamations",
    "telecommunication", "télécommunication", "professional", "comptabilité",
    "computer", "informatiques", "advertising", "publicité", "promotion",
    "credit card", "cartes de crédit", "bank", "banque", "frais",
    "cash", "transport", "office", "bureau", "travel", "déplacement",
    "management fee", "honoraires", "gestion", "incentives", "incitatif",
    "depreciation", "amortissement", "financial", "intérêts sur emprunts",
    "co-ownership", "copropriété", "copropriete", "shuttle", "navettes",
    "dues", "cotisations", "meal", "repas", "représentation",
    "bad debts", "mauvaises créances", "penalties", "allocation",
    "plate search", "collection fees", "software", "parking improvement",
    "new project", "indigo maintenance", "vinci", "ideal",
]

VALIDATION_SEARCH = {
    "total revenue": "_TOTAL_REVENUS_",
    "total revenus": "_TOTAL_REVENUS_",
    "total des revenus": "_TOTAL_REVENUS_",
    "total operation expenses": "_TOTAL_EXPENSES_",
    "total operating expenses": "_TOTAL_EXPENSES_",
    "total des frais d'exploitation": "_TOTAL_EXPENSES_",
    "operation surplus": "_OPERATION_SURPLUS_",
    "operating surplus": "_OPERATION_SURPLUS_",
    "net income": "_BENEFICE_NET_",
    "bénéfice net": "_BENEFICE_NET_",
    "benefice net": "_BENEFICE_NET_",
}

FULL_YEAR_MONTHS = {
    'January': 2, 'February': 3, 'March': 4, 'April': 5,
    'May': 6, 'June': 7, 'July': 8, 'August': 9,
    'September': 10, 'October': 11, 'November': 12, 'December': 13,
}

MONTH_MAP = {
    'janvier': 'January', 'février': 'February', 'fevrier': 'February',
    'mars': 'March', 'avril': 'April', 'mai': 'May', 'juin': 'June',
    'juillet': 'July', 'août': 'August', 'aout': 'August',
    'septembre': 'September', 'octobre': 'October',
    'novembre': 'November', 'décembre': 'December', 'decembre': 'December'
}

MONTH_COLUMN = {
    'January': 2, 'February': 3, 'March': 4, 'April': 5,
    'May': 6, 'June': 7, 'July': 8, 'August': 9,
    'September': 10, 'October': 11, 'November': 12, 'December': 13
}

# ============================================================================
# DETECT P&L TYPE
# ============================================================================

def detect_pnl_type(df):
    if len(df.columns) >= 14:
        data_cols = 0
        for col_idx in range(2, 14):
            if col_idx < len(df.columns):
                for row_idx in range(min(100, len(df))):
                    val = df.iloc[row_idx, col_idx - 1]
                    try:
                        if pd.notna(val) and float(val) != 0:
                            data_cols += 1
                            break
                    except:
                        pass
        if data_cols >= 3:
            return "full_year"
    return "single_month"

def detect_single_month_name(file_obj):
    if hasattr(file_obj, 'name'):
        for fr, en in MONTH_MAP.items():
            if fr in file_obj.name.lower():
                return en
    return None

# ============================================================================
# FIND YEAR TOTAL COLUMN
# ============================================================================

def find_year_total_column(df):
    for col_idx in range(len(df.columns)):
        for row_idx in range(min(15, len(df))):
            cell_val = str(df.iloc[row_idx, col_idx]).lower()
            if 'year total' in cell_val or 'année' in cell_val:
                return col_idx + 1
    for col_idx in range(len(df.columns) - 1, 1, -1):
        for row_idx in range(min(100, len(df))):
            val = df.iloc[row_idx, col_idx]
            try:
                if pd.notna(val) and float(val) != 0:
                    return col_idx + 1
            except:
                pass
    return len(df.columns)

# ============================================================================
# CHECK IF ROW IS A SUBTOTAL/TOTAL (should be skipped)
# ============================================================================

def is_excluded_row(account_name):
    """Check if this row is a subtotal, total, or header that should be skipped."""
    account_lower = account_name.lower().strip()
    for kw in EXCLUDE_KEYWORDS:
        if kw in account_lower:
            return True
    return False

def classify_account(account_name):
    """Determine if an unmatched account is revenue or expense."""
    account_lower = account_name.lower()
    for kw in REVENUE_KEYWORDS:
        if kw in account_lower:
            return "revenue"
    for kw in EXPENSE_KEYWORDS:
        if kw in account_lower:
            return "expense"
    return "unknown"

# ============================================================================
# EXCEL P&L EXTRACTION
# ============================================================================

def extract_from_full_year_pnl(df, debug_updates=None):
    data = {}
    unmatched_revenue = {}
    unmatched_expense = {}
    
    if debug_updates is not None:
        debug_updates.append("🔍 DEBUG - Unmatched rows:")
    
    for row_idx in range(len(df)):
        col_a = str(df.iloc[row_idx, 0]).strip().lower()
        if not col_a or col_a == 'nan':
            continue
        
        # Skip excluded rows (subtotals, totals, headers)
        if is_excluded_row(col_a):
            continue
        
        matched = False
        
        for search_term, template_row in ACCOUNT_SEARCH.items():
            if search_term in col_a:
                matched = True
                for month_name, col_idx in FULL_YEAR_MONTHS.items():
                    if col_idx < len(df.columns):
                        val = df.iloc[row_idx, col_idx - 1]
                        try:
                            amount = float(val) if pd.notna(val) else 0.0
                        except (ValueError, TypeError):
                            amount = 0.0
                        if month_name not in data:
                            data[month_name] = {}
                        if template_row not in data[month_name]:
                            data[month_name][template_row] = amount
                break
        
        if not matched:
            for search_term, validation_key in VALIDATION_SEARCH.items():
                if search_term in col_a:
                    matched = True
                    for month_name, col_idx in FULL_YEAR_MONTHS.items():
                        if col_idx < len(df.columns):
                            val = df.iloc[row_idx, col_idx - 1]
                            try:
                                amount = float(val) if pd.notna(val) else 0.0
                            except (ValueError, TypeError):
                                amount = 0.0
                            if month_name not in data:
                                data[month_name] = {}
                            if validation_key not in data[month_name]:
                                data[month_name][validation_key] = amount
                    break
        
        # Smart unmatched handling
        if not matched:
            has_value = False
            for month_name, col_idx in FULL_YEAR_MONTHS.items():
                if col_idx < len(df.columns):
                    val = df.iloc[row_idx, col_idx - 1]
                    try:
                        if pd.notna(val) and float(val) != 0:
                            has_value = True
                            break
                    except:
                        pass
            
            if has_value:
                account_type = classify_account(col_a)
                
                if debug_updates is not None:
                    jan_val = df.iloc[row_idx, 1] if len(df.columns) > 1 else 0
                    debug_updates.append(f"  ⚠️ UNMATCHED ({account_type}): Row {row_idx+1}: '{col_a[:60]}' = {jan_val}")
                
                for month_name, col_idx in FULL_YEAR_MONTHS.items():
                    if col_idx < len(df.columns):
                        val = df.iloc[row_idx, col_idx - 1]
                        try:
                            amount = float(val) if pd.notna(val) else 0.0
                        except (ValueError, TypeError):
                            amount = 0.0
                        
                        if amount != 0:
                            if account_type == "revenue":
                                if month_name not in unmatched_revenue:
                                    unmatched_revenue[month_name] = 0.0
                                unmatched_revenue[month_name] += amount
                            elif account_type == "expense":
                                if month_name not in unmatched_expense:
                                    unmatched_expense[month_name] = 0.0
                                unmatched_expense[month_name] += amount
    
    # Add unmatched revenue to Row 17 (Miscellaneous) - only if not already captured
    for month_name, amount in unmatched_revenue.items():
        if month_name not in data:
            data[month_name] = {}
        current = data[month_name].get(17, 0)
        data[month_name][17] = current + amount
        if debug_updates is not None:
            debug_updates.append(f"  ➕ Added {amount:,.2f} unmatched revenue to Row 17 for {month_name}")
    
    # Add unmatched expenses to Row 76
    for month_name, amount in unmatched_expense.items():
        if month_name not in data:
            data[month_name] = {}
        current = data[month_name].get(76, 0)
        data[month_name][76] = current + amount
        if debug_updates is not None:
            debug_updates.append(f"  ➖ Added {amount:,.2f} unmatched expense to Row 76 for {month_name}")
    
    return data

def extract_year_totals_from_pnl(df, debug_updates=None):
    year_col = find_year_total_column(df)
    
    if debug_updates is not None:
        debug_updates.append(f"📊 Budget Initial: Year Total column = {year_col}")
    
    data = {}
    unmatched_revenue = 0.0
    unmatched_expense = 0.0
    
    for row_idx in range(len(df)):
        col_a = str(df.iloc[row_idx, 0]).strip().lower()
        if not col_a or col_a == 'nan':
            continue
        
        # Skip excluded rows (subtotals, totals, headers)
        if is_excluded_row(col_a):
            continue
        
        matched = False
        
        for search_term, template_row in ACCOUNT_SEARCH.items():
            if search_term in col_a:
                matched = True
                if year_col <= len(df.columns):
                    val = df.iloc[row_idx, year_col - 1]
                    try:
                        amount = float(val) if pd.notna(val) else 0.0
                    except (ValueError, TypeError):
                        amount = 0.0
                    if template_row not in data:
                        data[template_row] = amount
                        if debug_updates is not None and amount != 0:
                            debug_updates.append(f"  ✅ Year Total: '{col_a[:50]}' → Row {template_row} = {amount:,.2f}")
                break
        
        if not matched:
            for search_term, validation_key in VALIDATION_SEARCH.items():
                if search_term in col_a:
                    matched = True
                    if year_col <= len(df.columns):
                        val = df.iloc[row_idx, year_col - 1]
                        try:
                            amount = float(val) if pd.notna(val) else 0.0
                        except (ValueError, TypeError):
                            amount = 0.0
                        if validation_key not in data:
                            data[validation_key] = amount
                    break
        
        # Smart unmatched for Year Total
        if not matched and year_col <= len(df.columns):
            val = df.iloc[row_idx, year_col - 1]
            try:
                amount = float(val) if pd.notna(val) else 0.0
            except (ValueError, TypeError):
                amount = 0.0
            
            if amount != 0:
                account_type = classify_account(col_a)
                if debug_updates is not None:
                    debug_updates.append(f"  ⚠️ UNMATCHED Year Total ({account_type}): '{col_a[:60]}' = {amount:,.2f}")
                
                if account_type == "revenue":
                    unmatched_revenue += amount
                elif account_type == "expense":
                    unmatched_expense += amount
    
    if unmatched_revenue != 0:
        data[17] = data.get(17, 0) + unmatched_revenue
        if debug_updates is not None:
            debug_updates.append(f"  ➕ Added {unmatched_revenue:,.2f} unmatched revenue to Row 17")
    
    if unmatched_expense != 0:
        data[76] = data.get(76, 0) + unmatched_expense
        if debug_updates is not None:
            debug_updates.append(f"  ➖ Added {unmatched_expense:,.2f} unmatched expense to Row 76")
    
    return data

def extract_from_single_month_pnl(df, month_name, debug_updates=None):
    data = {month_name: {}}
    data_col = None
    for col_idx in range(2, min(10, len(df.columns))):
        for row_idx in range(min(100, len(df))):
            val = df.iloc[row_idx, col_idx - 1]
            try:
                if pd.notna(val) and float(val) != 0:
                    data_col = col_idx
                    break
            except:
                pass
        if data_col:
            break
    if data_col is None:
        data_col = 2
    for row_idx in range(len(df)):
        col_a = str(df.iloc[row_idx, 0]).strip().lower()
        if not col_a or col_a == 'nan':
            continue
        if is_excluded_row(col_a):
            continue
        for search_term, template_row in ACCOUNT_SEARCH.items():
            if search_term in col_a:
                val = df.iloc[row_idx, data_col - 1]
                try:
                    amount = float(val) if pd.notna(val) else 0.0
                except (ValueError, TypeError):
                    amount = 0.0
                if template_row not in data[month_name]:
                    data[month_name][template_row] = amount
                break
        for search_term, validation_key in VALIDATION_SEARCH.items():
            if search_term in col_a:
                val = df.iloc[row_idx, data_col - 1]
                try:
                    amount = float(val) if pd.notna(val) else 0.0
                except (ValueError, TypeError):
                    amount = 0.0
                if validation_key not in data[month_name]:
                    data[month_name][validation_key] = amount
                break
    return data

def extract_from_pnl_excel(file_bytes, file_obj, parking_code, debug_updates=None):
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        tab_name = None
        for sn in xl.sheet_names:
            if parking_code.upper() in sn.upper():
                tab_name = sn
                break
        if tab_name is None:
            if debug_updates is not None:
                debug_updates.append(f"❌ Tab not found for {parking_code}")
            return {}
        if debug_updates is not None:
            debug_updates.append(f"📊 Found tab: {tab_name}")
        df = pd.read_excel(xl, sheet_name=tab_name, header=None)
        if debug_updates is not None:
            debug_updates.append(f"📐 P&L: {len(df)} rows x {len(df.columns)} cols")
        pnl_type = detect_pnl_type(df)
        if pnl_type == "full_year":
            if debug_updates is not None:
                debug_updates.append("📅 Full-year P&L detected")
            return extract_from_full_year_pnl(df, debug_updates)
        else:
            month_name = detect_single_month_name(file_obj)
            if month_name is None:
                month_name = 'January'
            if debug_updates is not None:
                debug_updates.append(f"📅 Single-month P&L: {month_name}")
            return extract_from_single_month_pnl(df, month_name, debug_updates)
    except Exception as e:
        if debug_updates is not None:
            debug_updates.append(f"❌ P&L error: {e}")
        return {}

# ============================================================================
# MAIN EXTRACTION
# ============================================================================

def extract_monthly_data(file_obj, parking_code, debug_updates=None):
    file_obj.seek(0)
    file_bytes = file_obj.read()
    if hasattr(file_obj, 'name') and file_obj.name.lower().endswith(('.xlsx', '.xls')):
        if debug_updates is not None:
            debug_updates.append("📊 Excel file detected")
        return extract_from_pnl_excel(file_bytes, file_obj, parking_code, debug_updates)
    if debug_updates is not None:
        debug_updates.append("❌ Not an Excel file")
    return {}

def extract_budget_initial_data(file_obj, parking_code, debug_updates=None):
    if file_obj is None:
        return None
    file_obj.seek(0)
    file_bytes = file_obj.read()
    if not hasattr(file_obj, 'name') or not file_obj.name.lower().endswith(('.xlsx', '.xls')):
        if debug_updates is not None:
            debug_updates.append("❌ Budget Initial: Not an Excel file")
        return None
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        tab_name = None
        for sn in xl.sheet_names:
            if parking_code.upper() in sn.upper():
                tab_name = sn
                break
        if tab_name is None:
            if debug_updates is not None:
                debug_updates.append(f"❌ Budget Initial: Tab not found for {parking_code}")
            return None
        df = pd.read_excel(xl, sheet_name=tab_name, header=None)
        if debug_updates is not None:
            debug_updates.append(f"📊 Budget Initial P&L: {len(df)} rows x {len(df.columns)} cols")
        return extract_year_totals_from_pnl(df, debug_updates)
    except Exception as e:
        if debug_updates is not None:
            debug_updates.append(f"❌ Budget Initial error: {e}")
        return None

# ============================================================================
# CHECK IF MONTH HAS REAL DATA
# ============================================================================

def month_has_data(month_data):
    for k, v in month_data.items():
        if not str(k).startswith('_') and v != 0:
            return True
    return False

# ============================================================================
# TEMPLATE FILLING
# ============================================================================

def fill_template(wb, all_data):
    sheet_name = None
    for sn in wb.sheetnames:
        if 'données' in sn.lower() or 'historique' in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        return ["❌ Sheet not found"]
    ws = wb[sheet_name]
    updates = []
    total = 0
    for month_en, month_data in all_data.items():
        col = MONTH_COLUMN.get(month_en)
        if col is None:
            continue
        col_letter = get_column_letter(col)
        month_cells = 0
        for key, amount in month_data.items():
            if str(key).startswith('_'):
                continue
            ws.cell(row=key, column=col).value = amount
            ws.cell(row=key, column=col).number_format = '#,##0.00 $'
            month_cells += 1
            total += 1
            updates.append(f"   ✅ {month_en} ({col_letter}{key}): {amount:,.2f} $")
        if month_cells > 0:
            updates.append(f"📊 {month_en}: {month_cells} cells")
    updates.append(f"\n📊 TOTAL: {total} yellow cells (formulas auto-calculate)")
    return updates

def fill_budget_initial(wb, budget_data):
    sheet_name = None
    for sn in wb.sheetnames:
        if 'budget' in sn.lower() and 'initial' in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        for sn in wb.sheetnames:
            if 'budget' in sn.lower():
                sheet_name = sn
                break
    if sheet_name is None:
        return ["❌ Budget Initial sheet not found"]
    ws = wb[sheet_name]
    updates = []
    total = 0
    col_s = 19
    for template_row, amount in budget_data.items():
        if str(template_row).startswith('_'):
            continue
        ws.cell(row=template_row, column=col_s).value = amount
        ws.cell(row=template_row, column=col_s).number_format = '#,##0.00 $'
        total += 1
        updates.append(f"   ✅ Budget Initial (S{template_row}): {amount:,.2f} $")
    updates.append(f"\n📊 Budget Initial: {total} cells filled in Column S")
    return updates

# ============================================================================
# ALLISON ANALYSIS
# ============================================================================

def allison_analyze(all_data, debug_updates=None):
    corrections = []
    for month_en, month_data in all_data.items():
        pnl_ben = month_data.get('_BENEFICE_NET_', 0)
        if pnl_ben == 0:
            continue
        pnl_rev = month_data.get('_TOTAL_REVENUS_', 0)
        pnl_exp = month_data.get('_TOTAL_EXPENSES_', 0)
        template_rev = sum(v for k, v in month_data.items() if k in [12, 13, 14, 15, 16, 17, 20, 22, 24])
        template_exp = sum(v for k, v in month_data.items() if k in [29, 30, 31, 32, 35, 36, 37, 38, 39, 40, 41, 42, 46, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 67, 68, 69, 70, 71, 72, 73, 74, 76])
        gap = pnl_ben - (template_rev - template_exp)
        if abs(gap) > 0.01:
            corrections.append(f"  🤖 Allison: {month_en} - Gap of {gap:,.2f} $")
    return corrections

# ============================================================================
# VALIDATION
# ============================================================================

def validate_results(wb, all_data):
    sheet_name = None
    for sn in wb.sheetnames:
        if 'données' in sn.lower() or 'historique' in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        return ["⚠️ Cannot validate"]
    results = []
    for month_en, month_data in all_data.items():
        ben = month_data.get('_BENEFICE_NET_')
        rev = month_data.get('_TOTAL_REVENUS_')
        exp = month_data.get('_TOTAL_EXPENSES_')
        results.append(f"\n📊 {month_en}:")
        if ben is not None and rev is not None and exp is not None:
            expected = rev - exp
            diff = abs(ben - expected)
            if diff > 0.01:
                results.append(f"   ⚠️ P&L Net: {ben:,.2f} $ | Expected: {expected:,.2f} $ (diff: {diff:,.2f} $)")
            else:
                results.append(f"   ✅ NET INCOME = {ben:,.2f} $")
        elif ben is not None:
            results.append(f"   ⚠️ Net: {ben:,.2f} $")
    return results

# ============================================================================
# APP.PY INTERFACE
# ============================================================================

def get_parking_codes_from_pnl(file_obj):
    codes = []
    if hasattr(file_obj, 'name'):
        file_obj.seek(0)
        if file_obj.name.lower().endswith(('.xlsx', '.xls')):
            try:
                xl = pd.ExcelFile(io.BytesIO(file_obj.read()))
                for sn in xl.sheet_names:
                    match = re.search(r'(CMO\d+|VMO\d+)', sn, re.IGNORECASE)
                    if match:
                        codes.append(match.group(1).upper())
            except:
                pass
        else:
            matches = re.findall(r'(CMO\d+)', file_obj.name, re.IGNORECASE)
            codes.extend(matches)
    file_obj.seek(0)
    return list(set([c.upper() for c in codes]))

def fix_excel(excel_file, monthly_files_current=None, monthly_files_previous=None,
              budget_initial_file=None, fiche_stationnement_file=None,
              parking_code=None, word_data=None):
    updates = []
    all_data = {}
    debug = []
    try:
        updates.append("=" * 60)
        updates.append("🏗️ BUDGET SYSTEM - WORKFLOW")
        updates.append("=" * 60)
        if isinstance(excel_file, bytes):
            wb = load_workbook(io.BytesIO(excel_file))
        else:
            excel_file.seek(0)
            wb = load_workbook(io.BytesIO(excel_file.read()))
        updates.append(f"✅ Template: {parking_code or 'Unknown'}")
        
        if monthly_files_previous:
            updates.append(f"\n📁 Previous year: {len(monthly_files_previous)} files...")
            for file_obj in monthly_files_previous:
                file_obj.seek(0)
                try:
                    debug.append(f"--- PREVIOUS: {getattr(file_obj, 'name', 'unknown')} ---")
                    data = extract_monthly_data(file_obj, parking_code, debug)
                    if data:
                        for mn, md in data.items():
                            all_data[mn] = md.copy()
                            cnt = len([k for k in md if not str(k).startswith('_')])
                            ben = md.get('_BENEFICE_NET_', 'N/A')
                            updates.append(f"   ✅ {mn}: {cnt} accounts | P&L Net: {ben} $")
                    else:
                        updates.append(f"   ❌ No data extracted")
                finally:
                    pass
        
        if monthly_files_current:
            updates.append(f"\n📁 Current year: {len(monthly_files_current)} files...")
            for file_obj in monthly_files_current:
                file_obj.seek(0)
                try:
                    debug.append(f"--- CURRENT: {getattr(file_obj, 'name', 'unknown')} ---")
                    data = extract_monthly_data(file_obj, parking_code, debug)
                    if data:
                        for mn, md in data.items():
                            if month_has_data(md):
                                all_data[mn] = md.copy()
                                cnt = len([k for k in md if not str(k).startswith('_')])
                                ben = md.get('_BENEFICE_NET_', 'N/A')
                                updates.append(f"   ✅ {mn}: {cnt} accounts | P&L Net: {ben} $")
                            else:
                                updates.append(f"   ⏭️ {mn}: Skipped (no data, keeping previous year)")
                    else:
                        updates.append(f"   ❌ No data extracted")
                finally:
                    pass
        
        if not all_data:
            updates.append("⚠️ No monthly data!")
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue(), updates
        
        if debug:
            updates.append("\n🔍 Debug:")
            updates.extend(debug)
        
        updates.append(f"\n📝 Filling Données historiques ({len(all_data)} months)...")
        updates.extend(fill_template(wb, all_data))
        
        if budget_initial_file:
            updates.append(f"\n📝 Processing Budget Initial...")
            budget_data = extract_budget_initial_data(budget_initial_file, parking_code, debug)
            if budget_data:
                updates.extend(fill_budget_initial(wb, budget_data))
            else:
                updates.append("   ⚠️ No Budget Initial data extracted")
        
        updates.append(f"\n🤖 Allison Analysis:")
        suggestions = allison_analyze(all_data, debug)
        if suggestions:
            updates.extend(suggestions)
        else:
            updates.append("   ✅ No gaps detected")
        updates.append(f"\n🔍 Validation:")
        updates.extend(validate_results(wb, all_data))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        updates.append("\n✅ DONE")
        return output.getvalue(), updates
    except Exception as e:
        updates.append(f"\n❌ {e}")
        import traceback
        updates.append(traceback.format_exc())
        return None, updates
