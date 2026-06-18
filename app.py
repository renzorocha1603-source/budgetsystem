# excel_fixer.py
import io
import re
import pandas as pd
import fitz
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import pdfplumber
import tempfile
import os

# ============================================================================
# CONFIGURATION
# ============================================================================

MONTHS_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

MONTH_NAMES_MAP = {
    "january": "January", "february": "February", "march": "March",
    "april": "April", "may": "May", "june": "June",
    "july": "July", "august": "August", "september": "September",
    "october": "October", "november": "November", "december": "December",
    "janvier": "January", "février": "February", "fevrier": "February",
    "mars": "March", "avril": "April", "mai": "May", "juin": "June",
    "juillet": "July", "août": "August", "aout": "August",
    "septembre": "September", "octobre": "October",
    "novembre": "November", "décembre": "December", "decembre": "December",
}

# ============================================================================
# PAGE 10 FRENCH LABELS -> STANDARD ENGLISH LABELS
# ============================================================================

PAGE10_FRENCH_LABELS = {
    "Revenus mensuels": "Monthly Revenues",
    "Revenus Journaliers": "Transient Revenue",
    "Revenus Lave-Auto": "Car-Wash Revenue",
    "Divers": "Miscellaneous",
    "(Gratuités - mensuels)": "Discount-Gratuities - Monthly",
    "TOTAL REVENUS": "TOTAL REVENUE",
    "Salaires Stationnement": "Parking wages",
    "Uniformes": "Uniforms",
    "Fourn. de stationnement": "Parking supplies",
    "Entretien réparation - Nettoyage": "R&M - Cleaning",
    "Entretien réparation - Equipement": "R&M - Equipement",
    "Entretien réparation - Général": "R&M - General",
    "Taxes et permis": "Tax & license",
    "Assurances Cautionnement": "Insurance & Guarantee",
    "Réclamations": "Claims",
    "Télécommunication": "Telecommunication",
    "Frais de cartes de crédit": "Credit Card fees",
    "Frais de bureau": "Office expenses",
    "Honoraires de gestion": "Percent Management fee",
    "BÉNÉFICE NET": "NET INCOME",
}

# ============================================================================
# YELLOW CELLS ONLY - Template Row Mapping (no formula rows!)
# ============================================================================

# These are the ONLY rows we write to (yellow user input cells)
# Formula rows: 18, 26, 33, 44, 47, 65, 82, 84, 86 - NEVER write to these!

TEMPLATE_YELLOW_ROWS = {
    # REVENUES (yellow rows 12-17, 20, 22, 24)
    "Transient Revenue": 12,
    "Monthly Revenues": 13,
    "Car-Wash Revenue": 14,
    "Hotel Revenue": 15,
    "Interests": 16,
    "Miscellaneous": 17,
    "Discount-Gratuities - Transient": 20,
    "Discount-Gratuities - Monthly": 22,
    
    # LABOUR (yellow rows 29-32)
    "Parking wages": 29,
    "Other wages": 30,
    "Training & Recr.": 31,
    "Uniforms": 32,
    
    # MAINTENANCE (yellow rows 35-43)
    "R&M - Cleaning": 35,
    "R&M - General": 36,
    "R&M - Equipement": 37,
    "R&M - Signs": 38,
    "R&M - Lines": 39,
    "Snow Removal": 40,
    "Parking supplies": 41,
    "Misc. Re-Billing": 42,
    
    # PUBLIC SERVICES (yellow row 46)
    "Public services": 46,
    
    # OVERHEAD (yellow rows 49-64)
    "Office expenses": 49,
    "Telecommunication": 50,
    "Rent": 51,
    "Travel expenses": 52,
    "Credit Card fees": 53,
    "Bank fees": 54,
    "Cash transportation fees": 55,
    "Claims": 56,
    "Insurance & Guarantee": 57,
    "Tax & license": 58,
    "Professional services": 59,
    "Equipment rent": 60,
    "Ad. & Promotion": 61,
    "Percent Management fee": 62,
    "Management Fees (Basic)": 63,
    "Incentives": 64,
    
    # OTHER EXPENSES (yellow rows 67-80)
    "Depreciation": 67,
    "Financial fees": 68,
    "Security": 69,
    "Co-ownership expenses": 70,
    "Shuttle expenses": 71,
    "Computer services": 72,
    "Bad debts": 73,
    "Dues & Subscription": 74,
    "Meal & Entertainment": 76,
}

# Month name -> Column in template (B=2, C=3, etc.)
MONTH_COLUMN = {
    'January': 2, 'February': 3, 'March': 4, 'April': 5,
    'May': 6, 'June': 7, 'July': 8, 'August': 9,
    'September': 10, 'October': 11, 'November': 12, 'December': 13
}

# ============================================================================
# PAGE 10 DETECTION (find the P&L page by content)
# ============================================================================

PAGE10_MARKERS = [
    "Revenus mensuels",
    "Revenus Journaliers", 
    "BÉNÉFICE NET",
    "Mois Courant"
]

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def safe_float(value, default=0.0):
    """Convert French Canadian number format to float."""
    try:
        if pd.isna(value) or value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').replace(' ', '').replace('\xa0', '')
            if value.startswith('(') and value.endswith(')'):
                return -safe_float(value[1:-1], default)
            if value.startswith('-'):
                return -safe_float(value[1:], default)
        return float(value)
    except (ValueError, TypeError):
        return default

def extract_month_from_filename(file_obj):
    """Extract English month name from filename."""
    if not hasattr(file_obj, 'name'):
        return None
    name = file_obj.name.lower()
    for key, value in MONTH_NAMES_MAP.items():
        if key in name:
            return value
    return None

def extract_year_from_filename(file_obj):
    """Extract year from filename."""
    if not hasattr(file_obj, 'name'):
        return None
    match = re.search(r'(20\d{2})', file_obj.name)
    if match:
        return int(match.group(1))
    return None

# ============================================================================
# PDF PAGE 10 EXTRACTION
# ============================================================================

def find_page10_in_pdf(doc):
    """Search ALL pages for the P&L page by content markers."""
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text()
        matches = sum(1 for m in PAGE10_MARKERS if m.lower() in text.lower())
        if matches >= 3:
            return page_num
    return None

def extract_page10_table(page):
    """Extract 28-row x 9-column table from P&L page using coordinates."""
    w = page.rect.width
    h = page.rect.height
    
    # Table area
    top = h * 0.06
    bottom = h * 0.95
    left = w * 0.03
    right = w * 0.97
    table_w = right - left
    table_h = bottom - top
    
    # Column widths
    col_widths = [0.25, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10]
    num_rows = 28
    row_h = table_h / num_rows
    
    data = []
    for row_idx in range(num_rows):
        row_data = []
        y = top + (row_idx * row_h)
        x = left
        for col_w in col_widths:
            cell_w = col_w * table_w
            rect = fitz.Rect(x + 1, y + 1, x + cell_w - 1, y + row_h - 1)
            text = page.get_text("text", clip=rect)
            row_data.append(' '.join(text.strip().split()))
            x += cell_w
        data.append(row_data)
    return data

def extract_from_digital_pdf(pdf_path):
    """Extract P&L data from digital PDF."""
    doc = fitz.open(pdf_path)
    page_num = find_page10_in_pdf(doc)
    
    if page_num is None:
        doc.close()
        return None
    
    page = doc[page_num]
    table = extract_page10_table(page)
    doc.close()
    
    # Parse table: match account names and get amounts from column 1 (Mois Courant)
    data = {}
    
    for row in table:
        account_text = row[0].strip() if len(row) > 0 else ""
        amount_text = row[1].strip() if len(row) > 1 else ""
        
        # Skip empty rows
        if not account_text:
            continue
        
        # Try to match account name to our labels
        for french_label, english_label in PAGE10_FRENCH_LABELS.items():
            if french_label.lower() in account_text.lower():
                amount = safe_float(amount_text)
                if amount != 0:
                    data[english_label] = amount
                break
    
    return data if len(data) >= 3 else None

# ============================================================================
# OCR FALLBACK FOR JANUARY (IMAGE PDF)
# ============================================================================

def extract_from_image_pdf(pdf_path):
    """Extract P&L data from image-based PDF using OCR."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return None
    
    doc = fitz.open(pdf_path)
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=300)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        
        try:
            text = pytesseract.image_to_string(img, lang='fra')
        except:
            try:
                text = pytesseract.image_to_string(img, lang='eng')
            except:
                text = pytesseract.image_to_string(img)
        
        # Check if this is the P&L page
        if "revenus mensuels" in text.lower() or "revenus journaliers" in text.lower():
            data = {}
            lines = text.split('\n')
            
            for line in lines:
                line = line.strip()
                if len(line) < 5:
                    continue
                
                # Try to match each French label
                for french_label, english_label in PAGE10_FRENCH_LABELS.items():
                    if english_label in data:
                        continue
                    
                    # Check if this line contains the account name
                    search_terms = french_label.lower().split()
                    if all(term in line.lower() for term in search_terms):
                        # Find a number in the line
                        numbers = re.findall(r'[\d\s,\.]+', line)
                        for num_str in numbers:
                            amount = safe_float(num_str)
                            if amount != 0:
                                data[english_label] = amount
                                break
                        break
            
            doc.close()
            return data if len(data) >= 3 else None
    
    doc.close()
    return None

# ============================================================================
# PDF FALLBACK WITH PDFPLUMBER
# ============================================================================

def extract_from_pdf_pdfplumber(pdf_path):
    """Extract using pdfplumber as last resort."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                
                if "revenus mensuels" in text.lower() or "revenus journaliers" in text.lower():
                    data = {}
                    lines = text.split('\n')
                    
                    for line in lines:
                        line = line.strip()
                        if len(line) < 5:
                            continue
                        
                        for french_label, english_label in PAGE10_FRENCH_LABELS.items():
                            if english_label in data:
                                continue
                            if french_label.lower() in line.lower():
                                numbers = re.findall(r'[\d\s,\.]+', line)
                                for num_str in numbers:
                                    amount = safe_float(num_str)
                                    if amount != 0:
                                        data[english_label] = amount
                                        break
                                break
                    
                    return data if len(data) >= 3 else None
    except Exception:
        pass
    
    return None

# ============================================================================
# MAIN EXTRACTION FUNCTION
# ============================================================================

def extract_from_pdf(pdf_path):
    """Extract P&L data from PDF - tries multiple methods."""
    # Method 1: Digital extraction (coordinates)
    data = extract_from_digital_pdf(pdf_path)
    if data and len(data) >= 3:
        return data
    
    # Method 2: OCR for image PDFs
    data = extract_from_image_pdf(pdf_path)
    if data and len(data) >= 3:
        return data
    
    # Method 3: pdfplumber fallback
    data = extract_from_pdf_pdfplumber(pdf_path)
    if data and len(data) >= 3:
        return data
    
    return None

# ============================================================================
# TEMPLATE FILLING (YELLOW CELLS ONLY)
# ============================================================================

def fill_template(wb, all_monthly_data):
    """Fill ONLY yellow cells in '2-Données historiques' sheet."""
    # Find the sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if 'données' in sn.lower() or 'historique' in sn.lower():
            sheet_name = sn
            break
    
    if sheet_name is None:
        return ["❌ Sheet '2-Données historiques' not found"]
    
    ws = wb[sheet_name]
    updates = []
    total_cells = 0
    
    for month_name, extracted_data in all_monthly_data.items():
        col = MONTH_COLUMN.get(month_name)
        if col is None:
            updates.append(f"⚠️ Unknown month: {month_name}")
            continue
        
        month_cells = 0
        for english_label, amount in extracted_data.items():
            # Skip totals and non-mapped labels
            if english_label in ['TOTAL REVENUE', 'NET INCOME', 'OPERATION SURPLUS', 'Total Operation expenses']:
                continue
            
            template_row = TEMPLATE_YELLOW_ROWS.get(english_label)
            if template_row is None:
                continue
            
            # Write to yellow cell
            cell = ws.cell(row=template_row, column=col)
            cell.value = amount
            cell.number_format = '#,##0.00'
            month_cells += 1
            total_cells += 1
            
            col_letter = get_column_letter(col)
            updates.append(f"   ✅ {month_name} ({col_letter}{template_row}): ${amount:,.2f} - {english_label}")
        
        if month_cells > 0:
            updates.append(f"📊 {month_name}: {month_cells} cells filled")
    
    updates.append(f"\n📊 TOTAL: {total_cells} yellow cells filled (formulas auto-calculate)")
    return updates

def validate_template(wb, all_monthly_data):
    """Validate NET INCOME from PDF matches REVENUS NETS in template."""
    sheet_name = None
    for sn in wb.sheetnames:
        if 'données' in sn.lower() or 'historique' in sn.lower():
            sheet_name = sn
            break
    
    if sheet_name is None:
        return ["⚠️ Cannot validate - sheet not found"]
    
    ws = wb[sheet_name]
    results = []
    
    for month_name, extracted_data in all_monthly_data.items():
        col = MONTH_COLUMN.get(month_name)
        if col is None:
            continue
        
        pdf_net_income = extracted_data.get('NET INCOME')
        template_net = ws.cell(row=86, column=col).value
        
        if pdf_net_income is not None and template_net is not None:
            diff = abs(pdf_net_income - template_net)
            if diff > 0.01:
                results.append(f"⚠️ {month_name}: PDF NET INCOME=${pdf_net_income:,.2f} ≠ Template REVENUS NETS=${template_net:,.2f}")
            else:
                results.append(f"✅ {month_name}: NET INCOME = REVENUS NETS = ${pdf_net_income:,.2f}")
        elif pdf_net_income is None:
            results.append(f"⚠️ {month_name}: NET INCOME not extracted from PDF")
    
    return results

# ============================================================================
# FUNCTIONS CALLED BY app.py
# ============================================================================

def get_parking_codes_from_pnl(file_obj):
    """Extract parking codes from uploaded file."""
    codes = []
    
    if hasattr(file_obj, 'name'):
        # From filename
        matches = re.findall(r'(CMO\d+)', file_obj.name, re.IGNORECASE)
        codes.extend(matches)
        
        # From PDF content
        file_obj.seek(0)
        if file_obj.name.lower().endswith('.pdf'):
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                    tmp.write(file_obj.read())
                    tmp_path = tmp.name
                
                doc = fitz.open(tmp_path)
                for page in doc:
                    text = page.get_text()
                    found = re.findall(r'(CMO\d+)', text, re.IGNORECASE)
                    codes.extend(found)
                doc.close()
                os.unlink(tmp_path)
            except Exception:
                pass
    
    file_obj.seek(0)
    return list(set([c.upper() for c in codes]))

def fix_excel(excel_file, monthly_files_current=None, monthly_files_previous=None,
              budget_initial_file=None, fiche_stationnement_file=None,
              parking_code=None, word_data=None):
    """
    MAIN FUNCTION - Extract P&L data from PDFs and fill template.
    Only writes to YELLOW cells. Formulas handle the rest.
    """
    updates = []
    all_monthly_data = {}
    
    try:
        updates.append("=" * 60)
        updates.append("🏗️ BUDGET SYSTEM - WORKFLOW")
        updates.append("=" * 60)
        
        # Load template
        if isinstance(excel_file, bytes):
            wb = load_workbook(io.BytesIO(excel_file))
        else:
            excel_file.seek(0)
            wb = load_workbook(io.BytesIO(excel_file.read()))
        
        updates.append(f"✅ Template loaded: {parking_code or 'Unknown'}")
        
        # Collect all monthly files
        all_files = []
        if monthly_files_current:
            all_files.extend(monthly_files_current)
        if monthly_files_previous:
            all_files.extend(monthly_files_previous)
        
        if not all_files:
            updates.append("\n⚠️ No monthly files provided!")
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue(), updates
        
        updates.append(f"\n📁 Processing {len(all_files)} files...")
        
        # Process each file
        for file_obj in all_files:
            month = extract_month_from_filename(file_obj)
            
            # Save to temp file
            file_obj.seek(0)
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                tmp.write(file_obj.read())
                tmp_path = tmp.name
            
            try:
                data = extract_from_pdf(tmp_path)
                
                if data and len(data) > 0:
                    all_monthly_data[month] = data
                    rev = data.get('TOTAL REVENUE', 'N/A')
                    net = data.get('NET INCOME', 'N/A')
                    updates.append(f"   ✅ {month}: {len(data)} accounts | Revenue: ${rev} | Net Income: ${net}")
                else:
                    updates.append(f"   ❌ {month}: No data extracted")
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        
        # Fill template
        if all_monthly_data:
            updates.append(f"\n📝 Filling YELLOW cells ({len(all_monthly_data)} months)...")
            fill_updates = fill_template(wb, all_monthly_data)
            updates.extend(fill_updates)
            
            # Validate
            updates.append(f"\n🔍 Validating NET INCOME = REVENUS NETS...")
            validations = validate_template(wb, all_monthly_data)
            updates.extend(validations)
        else:
            updates.append("\n⚠️ No data extracted from any file!")
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        updates.append("\n✅ WORKFLOW COMPLETE")
        return output.getvalue(), updates
        
    except Exception as e:
        updates.append(f"\n❌ ERROR: {str(e)}")
        import traceback
        updates.append(traceback.format_exc())
        return None, updates
