# excel_fixer.py - Converter + Allison + Regex Fallback
import io
import re
import pandas as pd
import fitz
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import tempfile
import os
from allison_extractor import call_allison

# ============================================================================
# PDF TO EXCEL CONVERTER (Claude's proven method)
# ============================================================================

PAGE10_EXCLUDE_KEYWORDS = [
    'CONCILIATION BI', 'ÉCARTS AU BUDGET', 'ECARTS AU BUDGET',
    'SECTION 1', 'SECTION 2', 'SECTION 3', 'SECTION 4',
    'MANUAL BILLING', 'FAITS SAILLANTS',
    'EXPLICATION DES ÉCARTS', 'AJUSTEMENT DÉPÔT',
    'CF. EXTRAIT BI', 'AVANT TAXES', "FRAIS D'OUVERTURE",
    'COMMENTAIRES', 'ANALYSE', 'SOMMAIRE',
    'MENSUELS TOTALES', 'JOURNALIERS TOTALES', 'DÉPENSES TOTALES',
]

def find_page10_in_pdf(doc):
    for page_num in range(len(doc)):
        page = doc[page_num]
        words = page.get_text("words")
        if not words or len(words) < 20:
            continue
        rows = {}
        for w in words:
            y_key = round(w[1] / 15) * 15
            if y_key not in rows:
                rows[y_key] = []
            rows[y_key].append(w[4])
        sorted_rows = sorted(rows.items())
        line_texts = [' '.join(row_words).upper() for _, row_words in sorted_rows]
        paired_lines = [line_texts[i] + ' ' + line_texts[i+1] for i in range(len(line_texts)-1)]
        combined_text = ' '.join(line_texts) + ' ' + ' '.join(paired_lines)
        excluded = False
        for kw in PAGE10_EXCLUDE_KEYWORDS:
            if kw in combined_text:
                excluded = True
                break
        if excluded:
            continue
        seq = ['REVENUS MENSUELS', 'REVENUS JOURNALIERS', 'REVENUS LAVE-AUTO']
        last_pos = -1
        seq_ok = True
        for term in seq:
            pos = combined_text.find(term)
            if pos == -1 or pos <= last_pos:
                seq_ok = False
                break
            last_pos = pos
        if not seq_ok:
            continue
        expense_found = any(term in combined_text for term in [
            'SALAIRES STATIONNEMENT', 'UNIFORMES', 'ENTRETIEN',
            'TAXES ET PERMIS', 'ASSURANCES', 'TÉLÉCOMMUNICATION'])
        if not expense_found:
            continue
        if 'TOTAL REVENUS' not in combined_text and 'TOTAL DES REVENUS' not in combined_text:
            continue
        if 'BÉNÉFICE NET' not in combined_text and 'BENEFICE NET' not in combined_text:
            continue
        return page_num
    return None

def convert_page10_to_excel(file_bytes):
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        page10_num = find_page10_in_pdf(doc)
        if page10_num is None:
            doc.close()
            return None
        page = doc[page10_num]
        words = page.get_text("words")
        if not words:
            doc.close()
            return None
        rows_dict = {}
        for word in words:
            x0, y0, x1, y1, text, block, line, word_no = word
            y_key = round(y0 / 10) * 10
            if y_key not in rows_dict:
                rows_dict[y_key] = []
            rows_dict[y_key].append((x0, text))
        sorted_y = sorted(rows_dict.keys())
        all_x = [x for y_key in sorted_y for x, _ in rows_dict[y_key]]
        if not all_x:
            doc.close()
            return None
        min_x = min(all_x)
        max_x = max(all_x)
        col_width = (max_x - min_x) / 9
        wb = Workbook()
        ws = wb.active
        ws.title = "Page10"
        headers = ["Account", "Mois Courant", "Budget période", "Écart Budget",
                   "An. Préc.", "Cumulatif courant", "Cumulatif budget",
                   "Écart Budget Cumul.", "An. Préc. Cumul."]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        excel_row = 2
        for y_key in sorted_y:
            row_items = rows_dict[y_key]
            row_items.sort(key=lambda item: item[0])
            cols = [''] * 9
            for x, text in row_items:
                col_idx = min(8, max(0, int((x - min_x) / col_width)))
                cols[col_idx] = (cols[col_idx] + ' ' + text).strip() if cols[col_idx] else text
            has_content = any(c.strip() for c in cols)
            if has_content:
                for col in range(9):
                    val = cols[col].strip()
                    if val:
                        try:
                            clean = val.replace('$', '').replace(',', '').replace(' ', '')
                            if clean.startswith('(') and clean.endswith(')'):
                                clean = '-' + clean[1:-1]
                            if clean.replace('.', '').replace('-', '').isdigit():
                                ws.cell(row=excel_row, column=col+1, value=float(clean))
                                ws.cell(row=excel_row, column=col+1).number_format = '#,##0.00'
                            else:
                                ws.cell(row=excel_row, column=col+1, value=val)
                        except:
                            ws.cell(row=excel_row, column=col+1, value=val)
                excel_row += 1
        doc.close()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except:
        return None

# ============================================================================
# CANADIAN FRENCH NUMBER PARSING (regex fallback)
# ============================================================================

def parse_amount(text):
    if not text:
        return None
    text = str(text).strip()
    is_negative = False
    if text.startswith('(') and text.endswith(')'):
        is_negative = True
        text = text[1:-1].strip()
    text = text.replace('$', '').strip()
    text = text.replace(' ', '').replace('\xa0', '').replace('\u202f', '')
    if text.startswith('-'):
        is_negative = True
        text = text[1:]
    if ',' in text:
        text = text.replace(',', '.')
    text = re.sub(r'[^\d\.\-]', '', text)
    try:
        value = float(text)
        return -value if is_negative else value
    except:
        return None

# ============================================================================
# REGEX SEARCH TERMS (fallback)
# ============================================================================

SEARCH_TERMS = {
    "revenus mensuels": 13, "revenus journaliers": 12, "revenus lave-auto": 14,
    "divers": 17, "gratuités - mensuels": 20, "salaires stationnement": 29,
    "uniformes": 32, "entretien réparation - nettoyage": 35,
    "entretien réparation - général": 36, "entretien réparation - equipement": 37,
    "fourn. de stationnement": 41, "frais de bureau": 49, "télécommunication": 50,
    "frais de cartes de crédit": 53, "réclamations": 56,
    "assurances cautionnement": 57, "taxes et permis": 58, "honoraires de gestion": 63,
}

VALIDATION_TERMS = {
    "bénéfice net": "_BENEFICE_NET_",
    "total revenus": "_TOTAL_REVENUS_",
    "total des frais d'exploitation": "_TOTAL_EXPENSES_",
}

# ============================================================================
# MONTH DETECTION
# ============================================================================

MONTH_MAP = {
    'janvier': 'January', 'février': 'February', 'fevrier': 'February',
    'mars': 'March', 'avril': 'April', 'mai': 'May', 'juin': 'June',
    'juillet': 'July', 'août': 'August', 'aout': 'August',
    'septembre': 'September', 'octobre': 'October',
    'novembre': 'November', 'décembre': 'December', 'decembre': 'December'
}

MONTH_COLUMN = {
    'January': 2, 'February': 3, 'March': 4, 'April': 5,
    'May': 6, 'June': 7, 'July': 8, 'August': 9,
    'September': 10, 'October': 11, 'November': 12, 'December': 13
}

# ============================================================================
# MAIN EXTRACTION - Converter → Allison → Regex Fallback
# ============================================================================

def extract_from_pdf(pdf_path, file_bytes=None, debug_updates=None):
    """Extract P&L data using the best available method."""
    
    if file_bytes is None:
        with open(pdf_path, 'rb') as f:
            file_bytes = f.read()
    
    # Method 1: Converter + Allison (structured data)
    if debug_updates is not None:
        debug_updates.append("🔍 Trying converter + Allison...")
    
    try:
        excel_output = convert_page10_to_excel(file_bytes)
        if excel_output:
            df = pd.read_excel(excel_output, sheet_name="Page10", engine='openpyxl')
            
            # Build structured text for Allison
            structured_lines = []
            for i in range(len(df)):
                row_parts = []
                for c in range(min(9, len(df.columns))):
                    val = df.iloc[i, c]
                    if pd.isna(val):
                        row_parts.append("nan")
                    else:
                        row_parts.append(str(val))
                structured_lines.append(" | ".join(row_parts))
            
            structured_text = "\n".join(structured_lines)
            
            if debug_updates is not None:
                debug_updates.append(f"📊 Converter produced {len(structured_lines)} rows for Allison")
            
            allison_result = call_allison(structured_text, debug_updates)
            
            if allison_result and "template_data" in allison_result:
                template_data = allison_result["template_data"]
                validation_data = allison_result.get("validation", {})
                
                data = {}
                for key_str, amount in template_data.items():
                    data[int(key_str)] = amount
                
                if "TOTAL_REVENUS" in validation_data:
                    data["_TOTAL_REVENUS_"] = validation_data["TOTAL_REVENUS"]
                if "TOTAL_EXPENSES" in validation_data:
                    data["_TOTAL_EXPENSES_"] = validation_data["TOTAL_EXPENSES"]
                if "BENEFICE_NET" in validation_data:
                    data["_BENEFICE_NET_"] = validation_data["BENEFICE_NET"]
                
                if len(data) >= 10:
                    return data
    except Exception as e:
        if debug_updates is not None:
            debug_updates.append(f"⚠️ Converter+Allison failed: {e}")
    
    # Method 2: Raw text + Allison (fallback)
    if debug_updates is not None:
        debug_updates.append("🔍 Trying raw text + Allison...")
    
    try:
        doc = fitz.open(pdf_path)
        pdf_text = ""
        for page_num in range(len(doc)):
            pdf_text += doc[page_num].get_text("text") + "\n"
        doc.close()
        
        allison_result = call_allison(pdf_text, debug_updates)
        
        if allison_result and "template_data" in allison_result:
            template_data = allison_result["template_data"]
            validation_data = allison_result.get("validation", {})
            
            data = {}
            for key_str, amount in template_data.items():
                data[int(key_str)] = amount
            
            if "TOTAL_REVENUS" in validation_data:
                data["_TOTAL_REVENUS_"] = validation_data["TOTAL_REVENUS"]
            if "TOTAL_EXPENSES" in validation_data:
                data["_TOTAL_EXPENSES_"] = validation_data["TOTAL_EXPENSES"]
            if "BENEFICE_NET" in validation_data:
                data["_BENEFICE_NET_"] = validation_data["BENEFICE_NET"]
            
            if len(data) >= 10:
                return data
    except Exception as e:
        if debug_updates is not None:
            debug_updates.append(f"⚠️ Raw text+Allison failed: {e}")
    
    # Method 3: Regex fallback
    if debug_updates is not None:
        debug_updates.append("🔍 Falling back to regex...")
    
    return extract_with_regex(pdf_path, debug_updates)

def extract_with_regex(pdf_path, debug_updates=None):
    doc = fitz.open(pdf_path)
    full_text = ""
    for page_num in range(len(doc)):
        full_text += doc[page_num].get_text("text") + "\n"
    doc.close()
    
    data = {}
    
    for search_term, template_row in SEARCH_TERMS.items():
        idx = full_text.lower().find(search_term)
        if idx == -1:
            continue
        after_text = full_text[idx + len(search_term):]
        match = re.search(r'(\d[\d\s]*,\d{2})\s*\$?', after_text[:30])
        if match:
            before = after_text[:match.start()].strip()
            is_neg = before.endswith('-') or before.endswith('(')
            amount = parse_amount(match.group(1))
            if amount is not None:
                if is_neg:
                    amount = -abs(amount)
                data[template_row] = amount
        else:
            data[template_row] = 0.0
    
    for search_term, validation_key in VALIDATION_TERMS.items():
        idx = full_text.lower().find(search_term)
        if idx == -1:
            continue
        after_text = full_text[idx + len(search_term):]
        match = re.search(r'(\d[\d\s]*,\d{2})\s*\$?', after_text[:30])
        if match:
            amount = parse_amount(match.group(1))
            if amount is not None:
                data[validation_key] = amount
    
    return data

# ============================================================================
# TEMPLATE FILLING
# ============================================================================

def fill_template(wb, all_data):
    sheet_name = None
    for sn in wb.sheetnames:
        if 'données' in sn.lower() or 'historique' in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        return ["❌ Sheet '2-Données historiques' not found"]
    
    ws = wb[sheet_name]
    updates = []
    total_cells = 0
    
    for month_en, month_data in all_data.items():
        col = MONTH_COLUMN.get(month_en)
        if col is None:
            updates.append(f"⚠️ Unknown month: {month_en}")
            continue
        
        col_letter = get_column_letter(col)
        month_cells = 0
        
        for key, amount in month_data.items():
            if str(key).startswith('_'):
                continue
            cell = ws.cell(row=key, column=col)
            cell.value = amount
            cell.number_format = '#,##0.00'
            month_cells += 1
            total_cells += 1
            updates.append(f"   ✅ {month_en} ({col_letter}{key}): ${amount:,.2f}")
        
        if month_cells > 0:
            updates.append(f"📊 {month_en}: {month_cells} cells filled")
    
    updates.append(f"\n📊 TOTAL: {total_cells} yellow cells filled (formulas auto-calculate)")
    return updates

# ============================================================================
# VALIDATION
# ============================================================================

def validate_results(wb, all_data):
    sheet_name = None
    for sn in wb.sheetnames:
        if 'données' in sn.lower() or 'historique' in sn.lower():
            sheet_name = sn
            break
    if sheet_name is None:
        return ["⚠️ Cannot validate - sheet not found"]
    
    results = []
    
    for month_en, month_data in all_data.items():
        col = MONTH_COLUMN.get(month_en)
        if col is None:
            continue
        
        pdf_ben = month_data.get('_BENEFICE_NET_')
        pdf_rev = month_data.get('_TOTAL_REVENUS_')
        pdf_exp = month_data.get('_TOTAL_EXPENSES_')
        
        results.append(f"\n📊 {month_en}:")
        
        if pdf_ben is not None and pdf_rev is not None and pdf_exp is not None:
            expected_net = pdf_rev - pdf_exp
            diff = abs(pdf_ben - expected_net)
            if diff > 0.01:
                results.append(f"   ⚠️ PDF BÉNÉFICE NET: ${pdf_ben:,.2f}")
                results.append(f"   📈 PDF Revenue: ${pdf_rev:,.2f} - PDF Expenses: ${pdf_exp:,.2f}")
                results.append(f"   📊 Expected Net: ${expected_net:,.2f} (diff: ${diff:,.2f})")
            else:
                results.append(f"   ✅ BÉNÉFICE NET = ${pdf_ben:,.2f}")
                results.append(f"   📈 Revenue: ${pdf_rev:,.2f} | Expenses: ${pdf_exp:,.2f} | Net: ${pdf_ben:,.2f}")
        elif pdf_ben is not None:
            results.append(f"   ⚠️ PDF BÉNÉFICE NET: ${pdf_ben:,.2f} (missing totals)")
        else:
            results.append(f"   ⚠️ BÉNÉFICE NET not found in PDF")
    
    return results

# ============================================================================
# APP.PY INTERFACE
# ============================================================================

def get_parking_codes_from_pnl(file_obj):
    codes = []
    if hasattr(file_obj, 'name'):
        matches = re.findall(r'(CMO\d+)', file_obj.name, re.IGNORECASE)
        codes.extend(matches)
    file_obj.seek(0)
    return list(set([c.upper() for c in codes]))

def fix_excel(excel_file, monthly_files_current=None, monthly_files_previous=None,
              budget_initial_file=None, fiche_stationnement_file=None,
              parking_code=None, word_data=None):
    updates = []
    all_data = {}
    debug_updates = []
    
    try:
        updates.append("=" * 60)
        updates.append("🏗️ BUDGET SYSTEM - WORKFLOW")
        updates.append("=" * 60)
        
        if isinstance(excel_file, bytes):
            wb = load_workbook(io.BytesIO(excel_file))
        else:
            excel_file.seek(0)
            wb = load_workbook(io.BytesIO(excel_file.read()))
        
        updates.append(f"✅ Template loaded: {parking_code or 'Unknown'}")
        
        all_files = []
        if monthly_files_current:
            all_files.extend(monthly_files_current)
        if monthly_files_previous:
            all_files.extend(monthly_files_previous)
        
        if not all_files:
            updates.append("⚠️ No files!")
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue(), updates
        
        updates.append(f"\n📁 Processing {len(all_files)} files...")
        
        for file_obj in all_files:
            month_en = None
            if hasattr(file_obj, 'name'):
                name_lower = file_obj.name.lower()
                for fr, en in MONTH_MAP.items():
                    if fr in name_lower:
                        month_en = en
                        break
            
            if month_en is None:
                updates.append(f"⚠️ Could not determine month for {getattr(file_obj, 'name', 'unknown')}")
                continue
            
            file_obj.seek(0)
            file_bytes = file_obj.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name
            
            try:
                debug_updates.append(f"--- {month_en} ---")
                data = extract_from_pdf(tmp_path, file_bytes, debug_updates)
                
                if data:
                    display_count = len([k for k in data if not str(k).startswith('_')])
                    all_data[month_en] = data
                    pdf_benefice = data.get('_BENEFICE_NET_', 'N/A')
                    updates.append(f"   ✅ {month_en}: {display_count} accounts | PDF BÉNÉFICE NET: ${pdf_benefice}")
                else:
                    updates.append(f"   ❌ {month_en}: No data extracted")
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        
        if debug_updates:
            updates.append("\n🔍 DEBUG:")
            updates.extend(debug_updates)
        
        if all_data:
            updates.append(f"\n📝 Filling YELLOW cells for {len(all_data)} months...")
            updates.append("   (Formula rows auto-calculate)")
            fill_updates = fill_template(wb, all_data)
            updates.extend(fill_updates)
            
            updates.append(f"\n🔍 Validation (PDF BÉNÉFICE NET vs Expected):")
            validations = validate_results(wb, all_data)
            updates.extend(validations)
        else:
            updates.append("\n⚠️ No data extracted from any file!")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        updates.append("\n✅ WORKFLOW COMPLETE")
        return output.getvalue(), updates
        
    except Exception as e:
        updates.append(f"\n❌ ERROR: {str(e)}")
        import traceback
        updates.append(traceback.format_exc())
        return None, updates
